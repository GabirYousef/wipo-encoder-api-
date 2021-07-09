import os
import openpyxl
import numpy as np
import torch
from PIL import Image
from torch import nn
from torch.nn import functional as F
from torch.utils.data import Dataset
from torch.utils.data.sampler import Sampler
from torchvision import transforms
from tqdm import tqdm   
import gc


class ImageReader(Dataset):
    def __init__(self, data_path, data_name, data_type, crop_type):
        if crop_type == 'cropped' and data_name not in ['car', 'cub']:
            raise NotImplementedError(
                'cropped data only works for car or cub dataset')

        data_dict = torch.load('{}/{}/{}_data_dicts.pth'.format(
            data_path, data_name, crop_type))[data_type]
        self.class_to_idx = dict(zip(sorted(data_dict), range(len(data_dict))))
        normalize = transforms.Normalize([0.485, 0.456, 0.406],
                                         [0.229, 0.224, 0.225])
        if data_type == 'train':
            self.transform = transforms.Compose([
                transforms.Resize((252, 252)),
                transforms.RandomCrop(224),
                transforms.RandomHorizontalFlip(),
                transforms.ToTensor(), normalize
            ])
        else:
            self.transform = transforms.Compose([
                transforms.Resize((224, 224)),
                transforms.ToTensor(), normalize
            ])
        self.images, self.labels = [], []
        for label, image_list in data_dict.items():
            self.images += image_list
            self.labels += [self.class_to_idx[label]] * len(image_list)

        self.data_name = data_name

    def __getitem__(self, index):
        path, target = self.images[index], self.labels[index]
        if self.data_name == 'rpc':
            path = path[29:]
        img = Image.open(path).convert('RGB')
        img = self.transform(img)
        return img, target

    def __len__(self):
        return len(self.images)


def recall(feature_vectors,
           feature_labels,
           rank,
           gallery_vectors=None,
           gallery_labels=None):
    num_features = len(feature_labels)
    feature_labels = torch.tensor(feature_labels,
                                  device=feature_vectors.device)
    gallery_vectors = feature_vectors if gallery_vectors is None else gallery_vectors

    dist_matrix = torch.cdist(feature_vectors.unsqueeze(0),
                              gallery_vectors.unsqueeze(0)).squeeze(0)

    if gallery_labels is None:
        dist_matrix.fill_diagonal_(float('inf'))
        gallery_labels = feature_labels
    else:
        gallery_labels = torch.tensor(gallery_labels,
                                      device=feature_vectors.device)

    idx = dist_matrix.topk(k=rank[-1], dim=-1, largest=False)[1]
    acc_list = []
    for r in rank:
        correct = (gallery_labels[idx[:, 0:r]] == feature_labels.unsqueeze(
            dim=-1)).any(dim=-1).float()
        acc_list.append((torch.sum(correct) / num_features).item())

    return acc_list


def recall_(feature_vectors,
            feature_labels,
            rank,
            gallery_vectors=None,
            gallery_labels=None):
    num_features = len(feature_labels)
    feature_labels = torch.tensor(feature_labels,
                                  device=feature_vectors.device)
    gallery_vectors = feature_vectors if gallery_vectors is None else gallery_vectors

    # split the test data into 300 chunks and loop over them
    # gc.collect()
    # torch.cuda.empty_cache()
    # dist_list = []
    # chunk_size = 300
    # for feature_vectors_chunk in torch.chunk(feature_vectors, chunk_size):
    #     dist_matrix_chunk = torch.cdist(feature_vectors_chunk.unsqueeze(0), gallery_vectors.unsqueeze(0).squeeze(0).detach()

    #     # dist_list.append(dist_matrix_chunk)

    #     # gc.collect()
    #     # torch.cuda.empty_cache()

    # dist_matrix=torch.cat(dist_list)

    if gallery_labels is None:
        dist_matrix.fill_diagonal_(float('inf'))
        gallery_labels = feature_labels
    else:
        gallery_labels = torch.tensor(gallery_labels,
                                      device=feature_vectors.device)

    idx = dist_matrix.topk(k=rank[-1], dim=-1, largest=False)[1]
    acc_list = []
    for r in rank:
        correct = (gallery_labels[idx[:, 0:r]] == feature_labels.unsqueeze(
            dim=-1)).any(dim=-1).float()
        acc_list.append((torch.sum(correct) / num_features).item())

    return acc_list


class LabelSmoothingCrossEntropyLoss(nn.Module):
    def __init__(self, smoothing=0.1, temperature=1.0):
        super().__init__()
        self.smoothing = smoothing
        self.temperature = temperature

    def forward(self, x, target):
        log_probs = F.log_softmax(x / self.temperature, dim=-1)
        nll_loss = -log_probs.gather(
            dim=-1, index=target.unsqueeze(dim=-1)).squeeze(dim=-1)
        smooth_loss = -log_probs.mean(dim=-1)
        loss = (1.0 - self.smoothing) * nll_loss + self.smoothing * smooth_loss
        return loss.mean()


class BatchHardTripletLoss(nn.Module):
    def __init__(self, margin=1.0):
        super().__init__()
        self.margin = margin

    @staticmethod
    def get_anchor_positive_triplet_mask(target):
        mask = torch.eq(target.unsqueeze(0), target.unsqueeze(1))
        mask.fill_diagonal_(False)
        return mask

    @staticmethod
    def get_anchor_negative_triplet_mask(target):
        labels_equal = torch.eq(target.unsqueeze(0), target.unsqueeze(1))
        mask = ~labels_equal
        return mask

    def forward(self, x, target):
        pairwise_dist = torch.cdist(x.unsqueeze(0), x.unsqueeze(0)).squeeze(0)

        mask_anchor_positive = self.get_anchor_positive_triplet_mask(target)
        anchor_positive_dist = mask_anchor_positive.float() * pairwise_dist
        hardest_positive_dist = anchor_positive_dist.max(1, True)[0]

        mask_anchor_negative = self.get_anchor_negative_triplet_mask(target)
        # make positive and anchor to be exclusive through maximizing the dist
        max_anchor_negative_dist = pairwise_dist.max(1, True)[0]
        anchor_negative_dist=pairwise_dist + max_anchor_negative_dist * \
            (1.0 - mask_anchor_negative.float())
        hardest_negative_dist = anchor_negative_dist.min(1, True)[0]

        loss = (F.relu(hardest_positive_dist - hardest_negative_dist +
                       self.margin))
        return loss.mean()


class MPerClassSampler(Sampler):
    def __init__(self, labels, batch_size, m=4):
        self.labels = np.array(labels)
        self.labels_unique = np.unique(labels)
        self.batch_size = batch_size
        self.m = m
        assert batch_size % m == 0, 'batch size must be divided by m'

    def __len__(self):
        return len(self.labels) // self.batch_size

    def __iter__(self):
        for _ in range(self.__len__()):
            labels_in_batch = set()
            inds = np.array([], dtype=np.int)

            while inds.shape[0] < self.batch_size:
                sample_label = np.random.choice(self.labels_unique)
                if sample_label in labels_in_batch:
                    continue

                labels_in_batch.add(sample_label)
                sample_label_ids = np.argwhere(
                    np.in1d(self.labels, sample_label)).reshape(-1)
                subsample = np.random.permutation(sample_label_ids)[:self.m]
                inds = np.append(inds, subsample)

            inds = inds[:self.batch_size]
            inds = np.random.permutation(inds)
            yield list(inds)


def combine_features(features_dir):
    all_db_images = []
    all_db_features = []
    for f in os.listdir(features_dir):
        db_dict = torch.load(os.path.join(features_dir, f))
        db_features = db_dict['db_features'].tolist
        db_images = db_dict['db_images']
        
        print(type(db_features))

        all_db_features += db_features
        all_db_images += db_images

    all_db_features = torch.tensor(all_db_features)
    out = {'db_images': all_db_images, 'db_features': all_db_features}
    torch.save(out, "db.pth")


def get_serial_numbers(excel_file):
    wb_obj = openpyxl.load_workbook(excel_file)
    sheet_obj = wb_obj.active

    serial_numbers = []

    for r in tqdm(range(sheet_obj.max_row - 1)):
        serial_numbers.append(sheet_obj.cell(row=r + 2, column=1).value)

    return serial_numbers


if __name__ == "__main__":
    serials = get_serial_numbers('assets/USPTO_PART_1.xlsx')
    print(serials[0])












